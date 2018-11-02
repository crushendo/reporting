import openpyxl
import os
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
import datetime
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis
#import mysql.connector

class scoutingReport():
    def sql2xl(self, startDateInput, endDateInput):
        #startDateInput = raw_input('What is the start date? (MM/DD/YYY)\n')
        startDate = datetime.datetime.strptime(str(startDateInput), '%m/%d/%Y')
        #endDateInput = raw_input('What is the end date?\n')
        endDate = datetime.datetime.strptime(str(endDateInput), '%m/%d/%Y')
        print(startDate)
        print(endDate)
        cwd = os.getcwd()
        print(cwd)

        # Establishing a connection to the mySQL database
        cnx = mysql.connector.connect(user='root', password='Kh18riku!',
                                      host='127.0.0.1',
                                      database='scouter',
                                      charset = 'utf8',
                                      use_unicode = 'FALSE')

        # Dynamically setting up the field columns and 'Mature' sheet
        mature_fields = cnx.cursor(dictionary=True)
        query = ("SELECT * FROM scoutapp_field WHERE age = 'Mature'")
        mature_fields.execute(query)
<<<<<<< HEAD
        report_path = "/Users/Ryan/projects18/reporting/scoutapp/utils/Scouting Report.xlsx"
=======
	
        report_path = "/home/lbadmin/projects18/reporting/scoutapp/utils/Scouting-Report.xlsx"
>>>>>>> 82798abc753ee3c924b11d929687c3f0a9979663
        report_wb = openpyxl.load_workbook(report_path)
        mature_sheet = report_wb.get_sheet_by_name('Mature Data')

        # Clearing old content from 'Mature' sheet
        i = 7
        while i < 500:
            highest_column = mature_sheet.max_column
            j = 1
            while j < highest_column:
                mature_sheet.cell(column=j,row=i).value = None
                j += 1
            i += 1
            j = 1

        # Populating 'Mature' sheet with field data
        i = 1
        for row in mature_fields:
            mature_sheet['A' + str(i + 6)].value = row["fieldName"]
            mature_sheet['B' + str(i + 6)].value = row["variety"]
            i += 1

        #Connecting to mysql db
        mature_data = cnx.cursor(dictionary=True)
        query = ("SELECT * FROM scoutapp_labelledata WHERE Age = 'Mature'")
        mature_data.execute(query)

        #Filling in scouting data
        for row in mature_data:
            i = 7
            report_last_row = int(mature_sheet.max_row)
            while i < report_last_row:
                if mature_sheet['A' + str(i)].value == row["Field"]:
                    #TODO: if date value for row in db fits the report date range
                    print(row["date"])
                    currentDate = datetime.datetime.strptime(row["date"], '%Y-%m-%d')
                    if currentDate >= startDate and currentDate <= endDate:
                        mature_sheet['D' + str(i)].value = row["date"]
                        if row["stop"]== "NW":
                            mature_sheet['I' + str(i)].value = int(row["adult"])
                            mature_sheet['O' + str(i)].value = int(row["eggs"])
                            mature_sheet['U' + str(i)].value = int(row["tapped"])
                            mature_sheet['AA' + str(i)].value = int(row["flush"])
                            mature_sheet['AG' + str(i)].value = row["LM"]
                            mature_sheet['AM' + str(i)].value = row["OD"]
                            mature_sheet['AS' + str(i)].value = row["SM"]
                        if row["stop"] == "SW":
                            mature_sheet['H' + str(i)].value = int(row["adult"])
                            mature_sheet['N' + str(i)].value = int(row["eggs"])
                            mature_sheet['T' + str(i)].value = int(row["tapped"])
                            mature_sheet['Z' + str(i)].value = int(row["flush"])
                            mature_sheet['AF' + str(i)].value = row["LM"]
                            mature_sheet['AL' + str(i)].value = row["OD"]
                            mature_sheet['AR' + str(i)].value = row["SM"]
                        if row["stop"] == "C":
                            mature_sheet['G' + str(i)].value = int(row["adult"])
                            mature_sheet['M' + str(i)].value = int(row["eggs"])
                            mature_sheet['S' + str(i)].value = int(row["tapped"])
                            mature_sheet['Y' + str(i)].value = int(row["flush"])
                            mature_sheet['AE' + str(i)].value = row["LM"]
                            mature_sheet['AK' + str(i)].value = row["OD"]
                            mature_sheet['AQ' + str(i)].value = row["SM"]
                        if row["stop"] == "NE":
                            mature_sheet['F' + str(i)].value = int(row["adult"])
                            mature_sheet['L' + str(i)].value = int(row["eggs"])
                            mature_sheet['R' + str(i)].value = int(row["tapped"])
                            mature_sheet['X' + str(i)].value = int(row["flush"])
                            mature_sheet['AD' + str(i)].value = row["LM"]
                            mature_sheet['AJ' + str(i)].value = row["OD"]
                            mature_sheet['AP' + str(i)].value = row["SM"]
                        if row["stop"] == "SE":
                            mature_sheet['E' + str(i)].value = int(row["adult"])
                            mature_sheet['K' + str(i)].value = int(row["eggs"])
                            mature_sheet['Q' + str(i)].value = int(row["tapped"])
                            mature_sheet['W' + str(i)].value = int(row["flush"])
                            mature_sheet['AC' + str(i)].value = row["LM"]
                            mature_sheet['AI' + str(i)].value = row["OD"]
                            mature_sheet['AO' + str(i)].value = row["SM"]
                i += 1

        # Dynamically setting up the field columns and 'Young' sheet
        young_fields = cnx.cursor(dictionary=True)
        young_query = ("SELECT * FROM scoutapp_field WHERE age = 'Young'")
        young_fields.execute(young_query)
        young_sheet = report_wb.get_sheet_by_name('Young Data')

        # Clearing old content from Young sheet
        i = 8
        while i < 500:
            highest_column = mature_sheet.max_column
            j = 1
            while j < highest_column:
                young_sheet.cell(column=j, row=i).value = None
                j += 1
            i += 1
            j = 1

        # Populating 'Young' sheet
        i = 1
        for row in young_fields:
            young_sheet['A' + str(i + 7)].value = row["fieldName"]
            young_sheet['B' + str(i + 7)].value = row["variety"]
            i += 1

        young_data = cnx.cursor(dictionary=True)
        # TODO: order by date, most recent last
        query = ("SELECT * FROM scoutapp_labelledata WHERE Age = 'Young'")
        young_data.execute(query)

        for row in young_data:
            i = 8
            report_last_row = int(young_sheet.max_row)
            while i < report_last_row:
                if young_sheet['A' + str(i)].value == row["Field"]:
                    currentDate = datetime.datetime.strptime(row["date"], '%Y-%m-%d')
                    if currentDate >= startDate and currentDate <= endDate:
                        young_sheet['D' + str(i)].value = row["date"]
                        if row["stop"] == "NW":
                            young_sheet['I' + str(i)].value = int(row["adult"])
                            young_sheet['O' + str(i)].value = int(row["eggs"])
                            young_sheet['U' + str(i)].value = int(row["Leafminer"])
                            young_sheet['AA' + str(i)].value = int(row["ODEggs"])
                            young_sheet['AG' + str(i)].value = row["ODLarva"]
                            young_sheet['AM' + str(i)].value = row["SpiderMites"]

                        if row["stop"] == "SW":
                            young_sheet['H' + str(i)].value = int(row["adult"])
                            young_sheet['N' + str(i)].value = int(row["eggs"])
                            young_sheet['T' + str(i)].value = int(row["Leafminer"])
                            young_sheet['Z' + str(i)].value = int(row["ODEggs"])
                            young_sheet['AF' + str(i)].value = row["ODLarva"]
                            young_sheet['AL' + str(i)].value = row["SpiderMites"]

                        if row["stop"] == "C":
                            young_sheet['G' + str(i)].value = int(row["adult"])
                            young_sheet['M' + str(i)].value = int(row["eggs"])
                            young_sheet['S' + str(i)].value = int(row["Leafminer"])
                            young_sheet['Y' + str(i)].value = int(row["ODEggs"])
                            young_sheet['AE' + str(i)].value = row["ODLarva"]
                            young_sheet['AK' + str(i)].value = row["SpiderMites"]

                        if row["stop"] == "NE":
                            young_sheet['F' + str(i)].value = int(row["adult"])
                            young_sheet['L' + str(i)].value = int(row["eggs"])
                            young_sheet['R' + str(i)].value = int(row["Leafminer"])
                            young_sheet['X' + str(i)].value = int(row["ODEggs"])
                            young_sheet['AD' + str(i)].value = row["ODLarva"]
                            young_sheet['AJ' + str(i)].value = row["SpiderMites"]

                        if row["stop"] == "SE":
                            young_sheet['E' + str(i)].value = int(row["adult"])
                            young_sheet['K' + str(i)].value = int(row["eggs"])
                            young_sheet['Q' + str(i)].value = int(row["Leafminer"])
                            young_sheet['W' + str(i)].value = int(row["ODEggs"])
                            young_sheet['AC' + str(i)].value = row["ODLarva"]
                            young_sheet['AI' + str(i)].value = row["SpiderMites"]

                i += 1

        report_wb.save('/home/lbadmin/projects18/reporting/scoutapp/utils/Scouting-Report-Temp.xlsx')

    def other_pests(self):
        report_path = "/home/lbadmin/projects18/reporting/scoutapp/utils/Scouting-Report-Temp.xlsx"
        report_wb = openpyxl.load_workbook(report_path)
        mature_sheet = report_wb.get_sheet_by_name('Mature Data')
        mature_report_sheet = report_wb.get_sheet_by_name('Mature Report')
        mature_standard_pests = ['LM', 'OD', 'SM']
        mature_max_row = mature_sheet.max_row
        print(mature_sheet.max_row)

        i = 7
        while i < mature_max_row:
            mature_report_sheet['AC' + str(i)].value = ''
            i += 1


        for pest in mature_standard_pests:
            print(pest)
            i = 7
            while i < mature_max_row:
                if pest == 'LM':
                    pest_columns = ['AC', 'AD', 'AE', 'AF', 'AG']
                elif pest == 'OD':
                    pest_columns = ['AI', 'AJ', 'AK', 'AL', 'AM']
                elif pest == 'SM':
                    pest_columns = ['AO', 'AP', 'AQ', 'AR', 'AS']
                j = 0
                pest_tally = 0
                while j < 5:
                    if mature_sheet[pest_columns[j] + str(i)].value == 'O':
                        pest_tally += 1
                    j += 1
                if pest_tally > 2:
                    print('got one: ' + str(i))
                    initial = mature_report_sheet['AC' + str(i)].value
                    print(type(initial))
                    if initial is None:
                        initial = ''
                    mature_report_sheet['AC' + str(i)].value = initial + ' ' + pest
                i += 1

        report_wb.save('/home/lbadmin/projects18/reporting/scoutapp/utils/Scouting-Report-Temp.xlsx')

    def other_pests(self):
        report_path = "/Users/Ryan/projects18/reporting/scoutapp/utils/Scouting Report.xlsx"
        report_wb = openpyxl.load_workbook(report_path)
        mature_sheet = report_wb.get_sheet_by_name('Mature Data')
        mature_report_sheet = report_wb.get_sheet_by_name('Mature Report')
        mature_standard_pests = ['LM', 'OD', 'SM']
        mature_max_row = mature_sheet.max_row
        print(mature_sheet.max_row)

        i = 7
        while i < mature_max_row:
            mature_report_sheet['AC' + str(i)].value = ''
            i += 1


        for pest in mature_standard_pests:
            print(pest)
            i = 7
            while i < mature_max_row:
                if pest == 'LM':
                    pest_columns = ['AC', 'AD', 'AE', 'AF', 'AG']
                elif pest == 'OD':
                    pest_columns = ['AI', 'AJ', 'AK', 'AL', 'AM']
                elif pest == 'SM':
                    pest_columns = ['AO', 'AP', 'AQ', 'AR', 'AS']
                j = 0
                pest_tally = 0
                while j < 5:
                    if mature_sheet[pest_columns[j] + str(i)].value == 'O':
                        pest_tally += 1
                    j += 1
                if pest_tally > 2:
                    print('got one: ' + str(i))
                    initial = mature_report_sheet['AC' + str(i)].value
                    print(type(initial))
                    if initial is None:
                        initial = ''
                    mature_report_sheet['AC' + str(i)].value = initial + ' ' + pest
                i += 1

        report_wb.save(report_path)







    def update_data(self):
<<<<<<< HEAD
        report_path = "/Users/Ryan/projects18/reporting/scoutapp/utils/Scouting Report.xlsx"
=======
        report_path = "/home/lbadmin/projects18/reporting/scoutapp/utils/Scouting-Report-Temp.xlsx"
>>>>>>> 82798abc753ee3c924b11d929687c3f0a9979663
        report_wb = openpyxl.load_workbook(report_path)
        mature_graph = report_wb.get_sheet_by_name('Mature Graph')

        i = 1
        while 1 == 1:
            date_iter = mature_graph['A' + str(i)].value
            if date_iter == None:
                break
            i += 1
        last_row = i - 1
        print('last row: ' + str(last_row))
        latest_date = mature_graph['A' + str(last_row)].value

        now = datetime.datetime.now()
        print(type(latest_date))
        today = now.strftime("%Y-%m-%d")
        today = datetime.datetime.strptime(today, '%Y-%m-%d')
        print(type(today))
        print('last row: ' + str(last_row))

        days_past = today - latest_date
        cycle_length = datetime.timedelta(days=14)
        print(days_past)
        # Add up ACP totals for the next 2 weeks (or more) from the db
        if days_past >= cycle_length:
            print("Days past = " + str(days_past))
            next_cycle_end = latest_date + cycle_length



    def create_graph(self):
<<<<<<< HEAD
        report_path = "/Users/Ryan/projects18/reporting/scoutapp/utils/Scouting Report.xlsx"
=======
        report_path = "/home/lbadmin/projects18/reporting/scoutapp/utils/Scouting-Report-Temp.xlsx"
>>>>>>> 82798abc753ee3c924b11d929687c3f0a9979663
        report_wb = openpyxl.load_workbook(report_path)
        mature_graph = report_wb.get_sheet_by_name('Mature Graph')
        last_row = int(mature_graph.max_row)
        data = Reference(mature_graph, min_col=2, min_row=1, max_col=4, max_row=last_row)

        c2 = LineChart()
        c2.title = "Mature Trees ACP"
        c2.style = 12
        c2.y_axis.title = "Psyllids"
        c2.y_axis.crossAx = 500
        c2.x_axis = DateAxis(crossAx=100)
        c2.x_axis.number_format = 'yyyy-mm-dd'
        c2.x_axis.majorTimeUnit = "days"
        c2.x_axis.title = "Date"

        c2.add_data(data, titles_from_data=True)
        dates = Reference(mature_graph, min_col=1, min_row=2, max_row=last_row)
        c2.set_categories(dates)
        s1 = c2.series[0]
        s1.marker.symbol = "square"
        s2 = c2.series[1]
        s2.marker.symbol = "square"
        s3 = c2.series[2]
        s3.marker.symbol = "square"

        mature_graph.add_chart(c2, "F1")

        report_wb.save(report_path)


a = scoutingReport()
#a.sql2xl()
#a.update_data()
a.other_pests()
#a.create_graph()
